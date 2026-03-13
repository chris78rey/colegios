from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook

from .models import WorkbookData


def load_workbook_data(path: Path) -> WorkbookData:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _load_csv(path)
    if suffix == ".xlsx":
        return _load_xlsx(path)
    raise ValueError(f"Archivo no soportado: {path.suffix}")


def _load_csv(path: Path) -> WorkbookData:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        raw_rows = list(reader)
    return _normalize_rows(path, raw_rows)


def _load_xlsx(path: Path) -> WorkbookData:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    raw_rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    return _normalize_rows(path, raw_rows)


def _normalize_rows(path: Path, raw_rows: list[list[object]]) -> WorkbookData:
    if not raw_rows:
        return WorkbookData(source_path=path, headers=[], rows=[])
    headers = [str(cell).strip() for cell in (raw_rows[0] or []) if str(cell).strip()]
    rows: list[dict[str, str]] = []
    for raw_row in raw_rows[1:]:
        row_dict: dict[str, str] = {}
        for index, header in enumerate(headers):
            value = raw_row[index] if index < len(raw_row) else ""
            row_dict[header] = "" if value is None else str(value).strip()
        if any(value for value in row_dict.values()):
            rows.append(row_dict)
    return WorkbookData(source_path=path, headers=headers, rows=rows)
